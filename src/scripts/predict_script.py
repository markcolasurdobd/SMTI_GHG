import src.predicting as pred
from src.data import PredictData
from src.datasets import make_2023
import pandas as pd

# Load data
d25 = PredictData()
d25.load_csv('./data/nan25.csv')
d25.transform()
X = d25.X

# Load model
model_path = r"C:\Users\10354191\OneDrive - BD\Projects\SMTI\GHG\models\model_24.pkl"
model = pred.load_model(model_path)

# Load vectorizer
vec_path = r"C:\Users\10354191\OneDrive - BD\Projects\SMTI\GHG\models\vec_24.pkl"
vectorizer = pred.load_vectorizer(vec_path)

# Make predictions
pred_df = pred.predict(model, vectorizer, X)

# from sklearn.metrics import classification_report
# print(classification_report(y, pred_df['preds']))


# sending to excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

red = PatternFill(fill_type = 'solid', start_color = 'FF0000', end_color = 'FF0000')
pred_df.to_excel('.test.xlsx', index = False)
wb = load_workbook('.test.xlsx')
ws = wb.active

for col in ws.iter_cols(min_row=2, min_col=2, max_col=2):
    for cell in col:
        if cell.value < 0.9:
            ws['A' + str(cell.row)].fill = red
            cell.fill = red



wb.save('.test.xlsx')
